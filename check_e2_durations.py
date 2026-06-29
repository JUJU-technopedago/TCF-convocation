import openpyxl

wb = openpyxl.load_workbook('JURYS FINAL TCF.xlsx')

sheets = ['TCF TP OBLIGATOIRE', 'TCF TP EE', 'TCF TP EO']

print("Vérification des cellules B2 (date) et E2 (durée) :")
print("=" * 60)

for sheet_name in sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        b2_value = ws['B2'].value
        e2_value = ws['E2'].value
        print(f"\n{sheet_name}:")
        print(f"  B2 (date)  : {b2_value}")
        print(f"  E2 (durée) : {e2_value}")
    else:
        print(f"\n{sheet_name}: non trouvé")
