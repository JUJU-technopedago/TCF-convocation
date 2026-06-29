#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pour ajouter SIANO Marco au fichier Excel et générer toutes les convocations
"""

import os
import sys
import pandas as pd
from datetime import datetime
import openpyxl
import shutil
from jury_excel_processor import JuryExcelProcessor
from pdf_generator import PDFGenerator

def add_siano_and_generate_all():
    """
    Ajoute SIANO Marco au fichier Excel et génère toutes les convocations
    """
    print("=" * 60)
    print("AJOUT DE SIANO MARCO ET GÉNÉRATION DE TOUTES LES CONVOCATIONS")
    print("=" * 60)
    
    # Trouver le fichier Excel le plus récent
    excel_files = [f for f in os.listdir('.') if f.startswith('juries_') and f.endswith('.xlsx')]
    
    if not excel_files:
        print("Aucun fichier Excel trouvé avec le format juries_*.xlsx")
        sys.exit(1)
    
    # Trier par date de modification (le plus récent en premier)
    excel_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    original_excel = excel_files[0]
    
    # Créer une copie du fichier Excel avec le numéro de version incrémenté
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_excel = f"juries_{timestamp}.xlsx"
    shutil.copy2(original_excel, new_excel)
    
    print(f"Fichier Excel original: {original_excel}")
    print(f"Nouvelle copie: {new_excel}")
    
    # Charger le workbook avec openpyxl
    print("Modification du fichier Excel pour ajouter SIANO Marco...")
    wb = openpyxl.load_workbook(new_excel)
    
    # Chercher la feuille B2
    b2_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Niveau B2':
            b2_sheet = wb[sheet_name]
            break
    
    if not b2_sheet:
        print("Onglet 'Niveau B2' non trouvé dans le fichier Excel")
        sys.exit(1)
    
    # Chercher la dernière ligne avec des données
    last_row = 0
    for row in range(1, 100):  # Limite arbitraire de 100 lignes
        if b2_sheet.cell(row=row, column=1).value is not None:
            last_row = row
    
    # Ajouter SIANO Marco à la ligne suivante
    row = last_row + 1
    b2_sheet.cell(row=row, column=1).value = "09:00"  # Heure de préparation (colonne A)
    b2_sheet.cell(row=row, column=2).value = "10:00"  # Heure de passage (colonne B)
    b2_sheet.cell(row=row, column=3).value = "032002032317"  # Numéro de candidat (colonne C)
    b2_sheet.cell(row=row, column=4).value = "SIANO Marco"  # Nom et prénom (colonne D)
    b2_sheet.cell(row=row, column=5).value = "15/07/1995"  # Date de naissance (colonne E)
    b2_sheet.cell(row=row, column=6).value = "marco.siano@example.com"  # Email (colonne F)
    b2_sheet.cell(row=row, column=7).value = "OUI"  # Besoins spéciaux (colonne G)
    
    # Sauvegarder le fichier Excel modifié
    wb.save(new_excel)
    print(f"SIANO Marco ajouté à l'onglet 'Niveau B2' du fichier {new_excel}")
    
    # Créer un dossier de sortie avec horodatage
    output_dir = f"output_all_avec_siano_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    
    print(f"Dossier de sortie: {output_dir}")
    print("-" * 60)
    
    # Initialiser le processeur Excel et le générateur PDF
    print("Initialisation du processeur Excel et du générateur PDF...")
    processor = JuryExcelProcessor(new_excel)
    
    # Définir les chemins
    template_path = "convocation_delf_template_modele.html"
    if not os.path.exists(template_path):
        template_path = "templates/convocation_delf_template_modele.html"
    
    # Vérifier les chemins des logos
    logo_af_path = "logoAF.svg"
    logo_delf_path = "logoDELF.svg"
    
    if not os.path.exists(logo_af_path):
        print(f"⚠️ Logo AF non trouvé à {logo_af_path}, utilisation du chemin par défaut")
        logo_af_path = "assets/logoAF.svg"
    
    if not os.path.exists(logo_delf_path):
        print(f"⚠️ Logo DELF non trouvé à {logo_delf_path}, utilisation du chemin par défaut")
        logo_delf_path = "assets/logoDELF.svg"
    
    # Initialiser le générateur PDF
    pdf_generator = PDFGenerator(
        excel_path=new_excel,
        template_path=template_path,
        logo_af_path=logo_af_path,
        logo_delf_path=logo_delf_path,
        output_dir=output_dir
    )
    
    # Charger les données des jurys
    print("Chargement des données du fichier Excel...")
    processor.load_jury_data()
    
    # Obtenir tous les candidats
    all_candidates = processor.get_all_candidates()
    
    if not all_candidates:
        print("Aucun candidat trouvé dans le fichier Excel.")
        return
    
    print(f"Nombre total de candidats trouvés: {len(all_candidates)}")
    
    # Compter les candidats par niveau
    niveau_counts = {}
    for candidat in all_candidates:
        niveau = candidat.get('niveau', 'Inconnu')
        niveau_counts[niveau] = niveau_counts.get(niveau, 0) + 1
    
    print("\nCandidats par niveau:")
    for niveau, count in sorted(niveau_counts.items()):
        print(f"  - Niveau {niveau}: {count} candidats")
    
    # Compter les candidats avec besoins spéciaux
    special_needs = [c for c in all_candidates if c.get('besoins_speciaux', False)]
    special_needs_count = len(special_needs)
    print(f"Candidats avec besoins spéciaux: {special_needs_count}")
    
    # Générer les PDF pour chaque candidat
    generated_files = []
    success_count = 0
    print("\nGénération des convocations PDF:")
    
    for i, candidat in enumerate(all_candidates):
        try:
            # Générer le PDF directement avec le processeur PDF
            pdf_path = pdf_generator.generate_pdf(candidat)
            
            generated_files.append(pdf_path)
            success_count += 1
            
            # Afficher la progression
            print(f"[{i+1}/{len(all_candidates)}] Généré: {os.path.basename(pdf_path)}")
            
            # Afficher des détails supplémentaires pour les candidats à besoins spéciaux
            if candidat.get('besoins_speciaux', False):
                print(f"  - Besoins spéciaux: Oui")
                print(f"  - Fin épreuve collective: {candidat.get('fin_ep_coll_affichage', '')}")
            
        except Exception as e:
            print(f"ERREUR pour le candidat {candidat.get('nom', '')} {candidat.get('prenom', '')}: {str(e)}")
    
    print("-" * 60)
    print(f"✅ {success_count} convocations générées avec succès dans le dossier {output_dir}")
    
    # Récapitulatif des besoins spéciaux
    if special_needs:
        print("\nCandidats avec besoins spéciaux détectés:")
        for candidate in special_needs:
            print(f"  - {candidate.get('nom', '')} {candidate.get('prenom', '')}: Niveau {candidate.get('niveau', '')}")
            print(f"    → Fin épreuve collective: {candidate.get('fin_ep_coll_affichage', candidate.get('fin_ep_coll', 'Non définie'))}")
            
    # Vérifier si SIANO Marco est dans la liste des candidats à besoins spéciaux
    siano = next((c for c in all_candidates if 'SIANO' in c.get('nom', '').upper() and 'Marco' in c.get('prenom', '')), None)
    if siano:
        print("\nInformations pour SIANO Marco:")
        print(f"  - Niveau: {siano.get('niveau', '')}")
        print(f"  - Besoins spéciaux: {siano.get('besoins_speciaux', False)}")
        print(f"  - Tiers-temps: {siano.get('tiers_temps', False)}")
        print(f"  - Fin épreuve collective: {siano.get('fin_ep_coll_affichage', siano.get('fin_ep_coll', 'Non définie'))}")
    else:
        print("\nSIANO Marco n'a pas été trouvé dans la liste des candidats.")
    
    # Exporter également vers Excel pour validation
    excel_output = os.path.join(output_dir, "candidats_export.xlsx")
    processor.export_to_standard_excel(excel_output)
    print(f"\nExport Excel créé: {excel_output}")
    
    print("\nTraitement terminé.")

if __name__ == "__main__":
    add_siano_and_generate_all()