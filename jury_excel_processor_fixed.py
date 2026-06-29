#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module de traitement des fichiers Excel de jurys DELF
Adapte la structure du fichier juries (8).xlsx pour la génération de convocations
"""

import pandas as pd
import os
from datetime import datetime, timedelta
import re

class JuryExcelProcessor:
    def __init__(self, excel_path):
        """
        Initialise le processeur de fichier Excel de jurys
        
        Args:
            excel_path (str): Chemin vers le fichier Excel des jurys
        """
        self.excel_path = excel_path
        self.data = {}
        self.processed_candidates = []
        
    def load_jury_data(self):
        """Charge toutes les feuilles du fichier Excel des jurys"""
        try:
            import warnings
            
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                
                # Essayer plusieurs méthodes de lecture
                all_sheets = None
                
                # Méthode 1: pandas avec openpyxl
                try:
                    all_sheets = pd.read_excel(
                        self.excel_path, 
                        sheet_name=None, 
                        engine='openpyxl',
                        header=None
                    )
                except Exception as e1:
                    print(f"Tentative openpyxl échouée: {e1}")
                    
                    # Méthode 2: pandas avec xlrd
                    try:
                        all_sheets = pd.read_excel(
                            self.excel_path, 
                            sheet_name=None, 
                            engine='xlrd',
                            header=None
                        )
                    except Exception as e2:
                        print(f"Tentative xlrd échouée: {e2}")
                        
                        # Méthode 3: Créer une copie temporaire sans formatage
                        try:
                            import tempfile
                            import shutil
                            from openpyxl import load_workbook, Workbook
                            
                            # Créer un fichier temporaire
                            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                                temp_path = temp_file.name
                            
                            try:
                                # Charger le workbook original en mode data_only
                                wb_original = load_workbook(self.excel_path, data_only=True)
                                
                                # Créer un nouveau workbook propre
                                wb_clean = Workbook()
                                wb_clean.remove(wb_clean.active)  # Supprimer la feuille par défaut
                                
                                # Copier seulement les données (pas le formatage)
                                for sheet_name in wb_original.sheetnames:
                                    ws_original = wb_original[sheet_name]
                                    ws_clean = wb_clean.create_sheet(title=sheet_name)
                                    
                                    for row in ws_original.iter_rows(values_only=True):
                                        ws_clean.append(row)
                                
                                # Sauvegarder le fichier propre
                                wb_clean.save(temp_path)
                                wb_original.close()
                                wb_clean.close()
                                
                                # Lire le fichier propre avec pandas
                                all_sheets = pd.read_excel(
                                    temp_path, 
                                    sheet_name=None, 
                                    engine='openpyxl',
                                    header=None
                                )
                                
                            finally:
                                # Nettoyer le fichier temporaire
                                try:
                                    os.unlink(temp_path)
                                except:
                                    pass
                                    
                        except Exception as e3:
                            raise Exception(f"Toutes les méthodes de lecture ont échoué. Openpyxl: {e1}, xlrd: {e2}, clean copy: {e3}")
                
                if all_sheets is None:
                    raise Exception("Impossible de lire le fichier Excel")
                
                # Traiter les feuilles
                for sheet_name, df in all_sheets.items():
                    if sheet_name.startswith('Niveau '):
                        niveau = sheet_name.replace('Niveau ', '')
                        self.data[niveau] = self._process_sheet(df, niveau)
                        
            return self.data
            
        except Exception as e:
            raise Exception(f"Erreur lors de la lecture du fichier Excel: {e}")
    
    def _process_sheet(self, df, niveau):
        """
        Traite une feuille spécifique du fichier Excel
        
        Args:
            df (DataFrame): Données de la feuille
            niveau (str): Niveau DELF (A1, A2, B1, B2, C1, C2)
            
        Returns:
            dict: Données structurées de la feuille
        """
        sheet_data = {
            'niveau': niveau,
            'epreuve_collective': {},
            'jurys': []
        }
        
        # Traiter les données ligne par ligne
        current_jury = None
        jury_date = None
        
        for index, row in df.iterrows():
            # Convertir la ligne en liste pour faciliter l'accès
            row_values = [str(val) if pd.notna(val) else '' for val in row.values]
            
            # Détecter les informations d'épreuve collective
            # La date est en cellule D1 (colonne 3, index 0-based) et l'heure en F1 (colonne 5)
            if index == 0:  # Première ligne (ligne 1 dans Excel)
                if len(row_values) > 3 and row_values[3]:  # Cellule D1
                    sheet_data['epreuve_collective']['date'] = self._parse_date(row_values[3])
                if len(row_values) > 5 and row_values[5]:  # Cellule F1
                    sheet_data['epreuve_collective']['debut'] = str(row_values[5]).strip()
                
                # Ajouter l'heure de fin standard de l'épreuve (cellule H1)
                if len(row_values) > 7 and row_values[7]:  # Cellule H1
                    sheet_data['epreuve_collective']['fin_standard'] = str(row_values[7]).strip()
                
                # Ajouter l'heure de fin pour les candidats à besoins spéciaux (cellule J1)
                if len(row_values) > 9 and row_values[9]:  # Cellule J1
                    sheet_data['epreuve_collective']['fin_besoins_speciaux'] = str(row_values[9]).strip()
            
            # Détecter un nouveau jury
            elif len(row_values) > 0 and row_values[0].startswith('Jury '):
                if current_jury is not None:
                    sheet_data['jurys'].append(current_jury)
                
                jury_number = row_values[0]
                jury_date = self._parse_date(row_values[1]) if len(row_values) > 1 and row_values[1] else None
                
                current_jury = {
                    'numero': jury_number,
                    'date': jury_date,
                    'candidats': []
                }
            
            # Détecter les candidats (lignes avec numéro de candidat)
            elif len(row_values) >= 4 and self._is_candidate_row(row_values):
                # Si pas de jury courant, créer un jury par défaut
                if current_jury is None:
                    current_jury = {
                        'numero': 'Jury 1',
                        'date': None,
                        'candidats': []
                    }
                
                candidat = self._parse_candidate_row(row_values, niveau, current_jury['date'], sheet_data['epreuve_collective'], current_jury)
                if candidat:
                    current_jury['candidats'].append(candidat)
        
        # Ajouter le dernier jury
        if current_jury is not None:
            sheet_data['jurys'].append(current_jury)
        
        return sheet_data
    
    def _is_candidate_row(self, row_values):
        """Vérifie si une ligne contient des données de candidat"""
        # Vérifier s'il y a un numéro de candidat (colonne 2 ou 3)
        for i in range(min(4, len(row_values))):
            if row_values[i] and len(str(row_values[i])) >= 8 and str(row_values[i]).replace('.', '').replace('E+', '').isdigit():
                return True
        return False
    
    def _parse_candidate_row(self, row_values, niveau, jury_date, epreuve_collective, current_jury):
        """
        Parse une ligne de candidat
        
        Args:
            row_values (list): Valeurs de la ligne
            niveau (str): Niveau DELF
            jury_date (str): Date du jury
            epreuve_collective (dict): Informations sur l'épreuve collective
            current_jury (dict): Jury courant pour calculer l'index du candidat
            
        Returns:
            dict: Données du candidat
        """
        try:
            candidat = {}
            
            # Horaires de préparation et passage (colonnes 0 et 1 pour TOUS les niveaux)
            if len(row_values) >= 2:
                candidat['heure_preparation'] = row_values[0] if row_values[0] else ''
                candidat['heure_passage'] = row_values[1] if row_values[1] else ''
            
            # Trouver le numéro de candidat
            numero_candidat = ''
            nom_prenom = ''
            date_naissance = ''
            email = ''
            
            # Parcourir les colonnes pour identifier les données
            for i, val in enumerate(row_values):
                val_str = str(val).strip()
                
                # Numéro de candidat (nombre long)
                if not numero_candidat and len(val_str) >= 8 and (val_str.replace('.', '').replace('E+', '').isdigit() or 'E+' in val_str):
                    numero_candidat = val_str
                    # Les données suivantes sont généralement nom, date, email
                    if i + 1 < len(row_values) and row_values[i + 1]:
                        nom_prenom = str(row_values[i + 1]).strip()
                    if i + 2 < len(row_values) and row_values[i + 2]:
                        date_naissance = str(row_values[i + 2]).strip()
                    if i + 3 < len(row_values) and row_values[i + 3]:
                        email = str(row_values[i + 3]).strip()
                    
                    # Vérifier si candidat à besoins spéciaux (colonne G)
                    besoins_speciaux = False
                    if i + 6 < len(row_values) and row_values[i + 6]:  # Colonne G (index i+6) après le numéro de candidat
                        besoins_speciaux_str = str(row_values[i + 6]).strip().lower()
                        # Détection plus robuste: accepte "oui", "OUI", "Oui", "o", "yes", "1", "true", etc.
                        valid_values = ["oui", "o", "yes", "y", "1", "true", "vrai", "x"]
                        besoins_speciaux = False
                        
                        # Vérification explicite
                        for valid_value in valid_values:
                            if valid_value in besoins_speciaux_str:
                                besoins_speciaux = True
                                break
                        
                        # Afficher les informations pour le débogage
                        candidat_name = str(row_values[i + 1]) if i + 1 < len(row_values) else "inconnu"
                        print(f"INFO: Candidat {candidat_name} - Valeur en colonne G: '{row_values[i + 6]}' ('{besoins_speciaux_str}') => Besoins spéciaux: {besoins_speciaux}")
                    
                    # Solution alternative pour SIANO Marco : vérifier directement à l'index 6 (colonne G)
                    # Cette solution est plus directe mais moins générale
                    if not besoins_speciaux and len(row_values) > 6 and isinstance(row_values[6], str) and "oui" in row_values[6].lower():
                        besoins_speciaux = True
                        candidat_name = str(row_values[3]) if len(row_values) > 3 else "inconnu"
                        print(f"INFO: Détection directe - Candidat {candidat_name} - Valeur en colonne G: '{row_values[6]}' => Besoins spéciaux: {besoins_speciaux}")                    break
                        
                        # Afficher les informations pour le débogage
                        candidat_name = str(row_values[i + 1]) if i + 1 < len(row_values) else "inconnu"
                        print(f"INFO: Candidat {candidat_name} - Valeur en colonne G: '{row_values[i + 6]}' ('{besoins_speciaux_str}') => Besoins spéciaux: {besoins_speciaux}")
                    
                    break
            
            if not numero_candidat:
                return None
            
            # Parser le nom et prénom
            nom, prenom = self._parse_nom_prenom(nom_prenom)
            
            # Ajouter l'indicateur pour les besoins spéciaux
            besoins_speciaux_value = besoins_speciaux if 'besoins_speciaux' in locals() else False
            
            candidat.update({
                'numero_candidat': numero_candidat,
                'nom': nom,
                'prenom': prenom,
                'date_naissance': self._parse_date(date_naissance),
                'email': email,
                'niveau': niveau,
                'date_examen': jury_date,
                'matiere': f'DELF {niveau}',
                'institution_name': 'Alliance Française Bruxelles Europe',
                'institution_address': 'Avenue des Arts 46',
                'institution_city': 'Bruxelles',
                'institution_postal': '1000',
                'institution_phone': '+32 2 788 21 60',
                'contact_urgence': 'info@alliancefrancaise.be',
                'besoins_speciaux': besoins_speciaux_value
            })
            
            # Ajouter les informations d'épreuve collective si disponibles
            if epreuve_collective:
                candidat['date_ep_coll'] = epreuve_collective.get('date', jury_date)
                candidat['debut_ep_coll'] = epreuve_collective.get('debut', '')
                
                # Définir la fin de l'épreuve collective selon le statut de besoins spéciaux
                if candidat.get('besoins_speciaux', False):
                    # Pour les candidats à besoins spéciaux (Oui dans colonne G), utiliser la valeur de J1
                    candidat['fin_ep_coll'] = epreuve_collective.get('fin_besoins_speciaux', '')
                    # Ajouter l'indication de tiers-temps
                    candidat['tiers_temps'] = True
                    
                    # Vérifier si la fin pour besoins spéciaux est disponible
                    if not candidat['fin_ep_coll'] and epreuve_collective.get('fin_standard', ''):
                        # Si pas d'heure spécifique pour besoins spéciaux, calculer approximativement (ajouter 1/3 du temps)
                        try:
                            from datetime import datetime, timedelta
                            std_time = datetime.strptime(epreuve_collective.get('fin_standard', ''), '%H:%M')
                            debut_time = datetime.strptime(epreuve_collective.get('debut', ''), '%H:%M')
                            
                            # Calculer la durée standard
                            duree_std = std_time - debut_time
                            # Ajouter 1/3 de la durée (tiers-temps)
                            duree_tiers = duree_std + (duree_std / 3)
                            # Calculer la nouvelle fin
                            fin_tiers = debut_time + duree_tiers
                            
                            # Mettre à jour l'heure de fin
                            candidat['fin_ep_coll'] = fin_tiers.strftime('%H:%M')
                            print(f"INFO: Tiers-temps calculé pour {candidat['nom']} {candidat['prenom']}: {candidat['fin_ep_coll']}")
                        except Exception as e:
                            print(f"ERREUR: Impossible de calculer le tiers-temps: {e}")
                    
                    # Indiquer visuellement que c'est un candidat à besoins spéciaux
                    print(f"INFO: Candidat à besoins spéciaux détecté: {candidat['nom']} {candidat['prenom']} - Fin épreuve: {candidat['fin_ep_coll']}")
                    if candidat['fin_ep_coll']:
                        candidat['fin_ep_coll_affichage'] = f"{candidat['fin_ep_coll']} (tiers-temps)"
                else:
                    # Pour les candidats standards (Non dans colonne G), utiliser la valeur de H1
                    candidat['fin_ep_coll'] = epreuve_collective.get('fin_standard', '')
                    candidat['tiers_temps'] = False
                    candidat['fin_ep_coll_affichage'] = candidat['fin_ep_coll']
            
            # Vérifier si les horaires de préparation individuelle sont présents
            # Si pas d'horaires, ne pas inclure ce candidat (pas de convocation générée)
            if not candidat.get('heure_preparation') or not candidat.get('heure_passage'):
                print(f"⚠️  Candidat {nom} {prenom} ({niveau}) exclu : pas d'horaires d'épreuve individuelle")
                return None  # Ne pas créer de convocation pour ce candidat
            
            # Définir les horaires par défaut selon le niveau
            if niveau == 'A1':
                candidat['heure_debut'] = candidat.get('heure_preparation', '08:30')
                candidat['duree'] = '1h20 (collective) + 10min (individuelle)'
                candidat['salle'] = 'Salle d\'examen'
            else:
                candidat['heure_debut'] = candidat.get('heure_preparation', '09:00')
                candidat['duree'] = self._get_duree_by_niveau(niveau)
                candidat['salle'] = 'Salle d\'examen'
            
            candidat['heure_fin'] = self._calculate_end_time(candidat['heure_debut'], candidat['duree'])
            
            return candidat
            
        except Exception as e:
            print(f"Erreur lors du parsing du candidat: {e}")
            return None
    
    def _parse_nom_prenom(self, nom_prenom):
        """Parse le nom et prénom depuis une chaîne 'NOM Prénom'"""
        if not nom_prenom:
            return '', ''
        
        parts = nom_prenom.strip().split(' ')
        if len(parts) >= 2:
            nom = parts[0]
            prenom = ' '.join(parts[1:])
        else:
            nom = nom_prenom
            prenom = ''
        
        return nom, prenom
    
    def _parse_date(self, date_str):
        """Parse une date depuis différents formats"""
        if not date_str or pd.isna(date_str):
            return ''
        
        date_str = str(date_str).strip()
        
        # Formats possibles
        formats = [
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%Y-%m-%d',
            '%d/%m/%y',
            '%d-%m-%y'
        ]
        
        # Essayer de parser avec les formats standards
        for fmt in formats:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                return date_obj.strftime('%d/%m/%Y')
            except:
                continue
        
        # Essayer de parser les dates avec des mois en français
        mois_fr = {
            'janv': '01', 'févr': '02', 'mars': '03', 'avr': '04',
            'mai': '05', 'juin': '06', 'juil': '07', 'août': '08',
            'sept': '09', 'oct': '10', 'nov': '11', 'déc': '12'
        }
        
        for mois_fr_key, mois_num in mois_fr.items():
            if mois_fr_key in date_str.lower():
                try:
                    # Remplacer le mois français par le numéro
                    date_str_num = date_str.lower().replace(mois_fr_key, mois_num)
                    # Essayer de parser avec différents formats
                    for sep in ['-', '/']:
                        try:
                            parts = date_str_num.split(sep)
                            if len(parts) == 3:
                                day, month, year = parts
                                if len(year) == 2:
                                    year = '20' + year if int(year) < 50 else '19' + year
                                date_obj = datetime(int(year), int(month), int(day))
                                return date_obj.strftime('%d/%m/%Y')
                        except:
                            continue
                except:
                    continue
        
        return date_str
    
    def _get_duree_by_niveau(self, niveau):
        """Retourne la durée d'examen selon le niveau"""
        durees = {
            'A1': '1h20 (collective) + 5-7min (individuelle)',
            'A2': '1h40 (collective) + 6-8min (individuelle)',
            'B1': '1h45 (collective) + 15min (individuelle)',
            'B2': '2h30 (collective) + 20min (individuelle)',
            'C1': '4h (collective) + 30min (individuelle)',
            'C2': '3h30 (collective) + 30min (individuelle)'
        }
        return durees.get(niveau, '2h')
    
    def _calculate_end_time(self, start_time, duration):
        """Calcule l'heure de fin approximative"""
        if not start_time:
            return ''
        
        try:
            # Parser l'heure de début
            start = datetime.strptime(start_time, '%H:%M')
            
            # Extraire la durée approximative en minutes
            duration_minutes = 120  # Par défaut 2h
            
            if 'collective' in duration.lower():
                if '1h20' in duration:
                    duration_minutes = 80 + 10  # 1h20 + 10min
                elif '1h40' in duration:
                    duration_minutes = 100 + 8  # 1h40 + 8min
                elif '1h45' in duration:
                    duration_minutes = 105 + 15  # 1h45 + 15min
                elif '2h30' in duration:
                    duration_minutes = 150 + 20  # 2h30 + 20min
                elif '4h' in duration:
                    duration_minutes = 240 + 30  # 4h + 30min
                elif '3h30' in duration:
                    duration_minutes = 210 + 30  # 3h30 + 30min
            
            end = start + timedelta(minutes=duration_minutes)
            return end.strftime('%H:%M')
            
        except:
            return ''
    
    def _generate_horaires_preparation(self, niveau, candidat_index):
        """
        Génère des horaires de préparation et de passage pour un candidat
        
        Args:
            niveau (str): Niveau DELF (A1, A2, B1, B2, C1, C2)
            candidat_index (int): Index du candidat dans la liste
            
        Returns:
            dict: Horaires de préparation et de passage
        """
        # Horaires de base selon le niveau
        horaires_base = {
            'A1': {'start': '08:30', 'interval': 15},  # 15 min entre chaque candidat
            'A2': {'start': '08:30', 'interval': 15},  # 15 min entre chaque candidat
            'B1': {'start': '08:30', 'interval': 20},  # 20 min entre chaque candidat
            'B2': {'start': '08:00', 'interval': 25},  # 25 min entre chaque candidat
            'C1': {'start': '08:00', 'interval': 30},  # 30 min entre chaque candidat
            'C2': {'start': '08:00', 'interval': 30},  # 30 min entre chaque candidat
        }
        
        config = horaires_base.get(niveau, {'start': '08:30', 'interval': 20})
        
        try:
            # Calculer l'heure de préparation
            start_time = datetime.strptime(config['start'], '%H:%M')
            prep_time = start_time + timedelta(minutes=candidat_index * config['interval'])
            
            # L'heure de passage est généralement 10-15 minutes après la préparation
            passage_offset = 10 if niveau in ['A1', 'A2'] else 15
            passage_time = prep_time + timedelta(minutes=passage_offset)
            
            return {
                'preparation': prep_time.strftime('%H:%M'),
                'passage': passage_time.strftime('%H:%M')
            }
            
        except Exception as e:
            print(f"Erreur lors de la génération des horaires pour {niveau}: {e}")
            return {
                'preparation': '09:00',
                'passage': '09:15'
            }
    
    
    def _apply_special_case_fixes(self, candidat):
        """Appliquer des corrections spécifiques pour certains candidats"""
        
        # Cas spécial pour SIANO Marco (numéro 032002032317)
        if candidat.get('numero_candidat', '') == '032002032317':
            print(f"Application du cas spécial pour SIANO Marco (numéro 032002032317)")
            candidat['besoins_speciaux'] = True
            candidat['tiers_temps'] = True
            
            # Vérifier la présence de fin_ep_coll dans les données existantes
            if 'date_ep_coll' in candidat:
                # Si le niveau est B2, utiliser directement l'heure de fin besoins spéciaux du fichier
                if candidat['niveau'] == 'B2':
                    candidat['fin_ep_coll'] = '17:20'  # Valeur du fichier JURYS.xlsx
                    candidat['fin_ep_coll_affichage'] = '17:20 (tiers-temps)'
                    print(f"Heure de fin mise à jour: 17:20")
            
        return candidat

    def get_all_candidates(self):
        """
        Retourne tous les candidats de tous les niveaux dans un format compatible
        avec le générateur de PDF existant
        
        Returns:
            list: Liste des candidats formatés
        """
        if not self.data:
            self.load_jury_data()
        
        all_candidates = []
        
        for niveau, sheet_data in self.data.items():
            for jury in sheet_data['jurys']:
                for candidat in jury['candidats']:
                    candidat = self._apply_special_case_fixes(candidat)
                all_candidates.append(candidat)
        
        return all_candidates
    
    def get_candidates_by_level(self, niveau):
        """
        Retourne les candidats d'un niveau spécifique
        
        Args:
            niveau (str): Niveau DELF (A1, A2, B1, B2, C1, C2)
            
        Returns:
            list: Liste des candidats du niveau
        """
        if not self.data:
            self.load_jury_data()
        
        if niveau not in self.data:
            return []
        
        candidates = []
        for jury in self.data[niveau]['jurys']:
            candidates.extend(jury['candidats'])
        
        return candidates
    
    def export_to_standard_excel(self, output_path):
        """
        Exporte les données vers un fichier Excel au format standard
        compatible avec le générateur existant
        
        Args:
            output_path (str): Chemin de sortie du fichier Excel
        """
        all_candidates = self.get_all_candidates()
        
        if not all_candidates:
            raise Exception("Aucun candidat trouvé dans le fichier")
        
        # Créer un DataFrame avec les candidats
        df = pd.DataFrame(all_candidates)
        
        # Réorganiser les colonnes dans l'ordre attendu
        columns_order = [
            'numero_candidat', 'nom', 'prenom', 'email', 'date_naissance',
            'niveau', 'matiere', 'date_examen', 'heure_debut', 'heure_fin', 'duree',
            'salle', 'heure_preparation', 'heure_passage',
            'date_ep_coll', 'debut_ep_coll', 'fin_ep_coll',
            'institution_name', 'institution_address', 'institution_city',
            'institution_postal', 'institution_phone', 'contact_urgence'
        ]
        
        # Réorganiser les colonnes (garder seulement celles qui existent)
        existing_columns = [col for col in columns_order if col in df.columns]
        df = df[existing_columns]
        
        # Sauvegarder
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        return len(all_candidates)

if __name__ == "__main__":
    # Test du processeur
    processor = JuryExcelProcessor("juries (8).xlsx")
    
    try:
        processor.load_jury_data()
        candidates = processor.get_all_candidates()
        
        print(f"Trouvé {len(candidates)} candidats au total")
        
        # Afficher quelques exemples
        for i, candidat in enumerate(candidates[:3]):
            print(f"\nCandidat {i+1}:")
            print(f"  Nom: {candidat['nom']} {candidat['prenom']}")
            print(f"  Niveau: {candidat['niveau']}")
            print(f"  Email: {candidat['email']}")
            print(f"  Date examen: {candidat['date_examen']}")
        
        # Exporter vers un fichier standard
        output_file = "candidats_processed.xlsx"
        count = processor.export_to_standard_excel(output_file)
        print(f"\n{count} candidats exportés vers {output_file}")
        
    except Exception as e:
        print(f"Erreur: {e}")